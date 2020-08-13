from fuzzywuzzy import fuzz
from openpyxl import load_workbook
import re
from collections import defaultdict
from colorama import Fore, Style
import ray
import psutil



# Benchmark 13/08/20 SERIAL /w ray initialised
# # Σ Addresses: 3046
# matched 2638 addresses
# # Σ unmatched addresses: 408
# Tier 2 matches =  390
# Tier 3 matches =  134
# Tier 1 matches =  2114
# python3 addr_match.py  103.43s user 4.82s system 104% cpu 1:43.79 total

# # Σ Addresses: 3046
# matched 2638 addresses
# # Σ unmatched addresses: 408
# Tier 2 matches =  390
# Tier 3 matches =  134
# Tier 1 matches =  2114
# python3 addr_match.py  6.49s user 2.14s system 28% cpu 30.252 total


# Data structure storing a dictionary.
# The keys of the dictionary are numerical
# where, the key 1 maps to a list of addr_match objects such that
# every addr_match object in this list stores an address and its
# matching address iff the match ratio is > 63 and tkn > 73
# The key 2 stores the exact same except matches have a ratio
# of > 53 and tkn > 63.
# Key 3 has the same but each ratio is 10 less than its predecessor
class match_structure:
    def __init__(self):
        self.match_d = defaultdict(list)

    # Function that given an addr_match object (fully initialised)
    # , will insert the addr_match into the match dictionary
    # generating the correct key
    def insert_match(self, addr_match):
        # tier will never be -1 since addr_match must be a potential match
        tier = chck_ratio(addr_match)
        # insert the match to the list at key: tier
        self.match_d[tier].append(addr_match)

    # prints a debug stat of the number of matches in each tier
    def print(self):
        for tier in self.match_d:
            print("Tier", tier, "matches =", set_console_col(tier),len(self.match_d[tier])*2, Style.RESET_ALL)


# class representing an address and its fuzzy matching ratio
# Intended purpose: for every address, store a list of addr_match objects
# ,of the matching addresses in the other dataset(s)
class addr_match:
    def __init__(self, addr, f_ratio, f_tkn_ratio, m_addr):
        self.addr = addr
        self.f_ratio = f_ratio
        self.f_tkn_ratio = f_tkn_ratio
        self.m_addr = m_addr


# Helper function that standardises a given string address
def standardise_addr(addr):
    # apply any transformations to addresses
    # strip all commas away from the address and make it uppercase
    addr = addr.replace(',', '').upper()
    # remove the word 'flat'
    addr = addr.replace('FLAT', '')
    # remove the word 'apartment'
    addr = addr.replace('APARTMENT', '')
    # remove the word 'unit'
    addr = addr.replace('UNIT', '')
    # remove the word 'England'
    addr = addr.replace('ENGLAND', '')
    # remove the word 'France'
    addr = addr.replace('FRANCE', '')
    # remove the word 'Germany'
    addr = addr.replace('GERMANY', '')
    # remove the word 'drive'
    # strip() removes leading and trailing whitespace
    addr = addr.strip()
    return addr


# helper function that determines if the addr_match object
# qualifies for a match
# function returns the tier (int) of the match type
def chck_ratio(addr_match):
    if addr_match.f_ratio >= 63 and addr_match.f_tkn_ratio >= 73:
        return 1
    elif addr_match.f_ratio >= 60 and addr_match.f_tkn_ratio >= 65:
        return 2
    elif addr_match.f_ratio >= 55 and addr_match.f_tkn_ratio >= 60:
        return 3
    else:
        # at this point the match is so bad, it should be discarded
        return -1


# sets the param colour based on the tier
def set_console_col(tier):
    colour = Fore.GREEN
    if tier == 2:
        colour = Fore.YELLOW
    elif tier == 3:
        colour = Fore.RED
    return colour


# Splits a given list into parts of approximately equal length
# src: https://stackoverflow.com/questions/2130016/splitting-a-list-into-n-parts-of-approximately-equal-length
def split_list(list, parts):
    k, m = divmod(len(list), parts)
    return (list[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(parts))


# Function that finds matching addresses in sa2_l (global) for addresses in
# sa1_l.
# sa1_l = parsed addresses from dataset 1 as a list
# sa2_l = parsed addresses from dataset 2 as a list
# sa1_l_s = subset of sa1_l
#   if number of cpu cores = n then sa1_l is split into n parts of approximate
#   equal length - sa1_l_s is one of those parts
# function returns list matched_addr containing addr_match objects
@ray.remote
def find_matches_parallel(sa1_l_s):
    # find matching addresses from each list
    matched_addr = []
    for addr_1 in sa1_l_s:
        # populate a list of matching addr_match objects for some address in list 1
        matches = []
        # iterate through every address in list 2
        mod_addr_2 = ''
        # standardise the address first
        mod_addr_1 = standardise_addr(addr_1)
        for addr_2 in sa2_l:
            # standardise the address first
            mod_addr_2 = standardise_addr(addr_2)

            # determine fuzzy matching ratio
            f_ratio = fuzz.ratio(mod_addr_1, mod_addr_2)
            f_tkn_ratio = fuzz.token_sort_ratio(mod_addr_1, mod_addr_2)
            # f_tkn_set_ratio = fuzz.token_set_ratio(mod_addr_1, mod_addr_2)
            # f_partial_ratio = fuzz.partial_ratio(mod_addr_1, mod_addr_2)

            # If the first part of the address is a number, check if they both match
            # E.g 1, Charter House and 2, Charter House should not match
            # At this point, the Address has been standardised
            # see standardise_addr() fn for more details
            # get first part of the address
            addr_num1 = mod_addr_1.split(' ')[0]
            addr_num2 = mod_addr_2.split(' ')[0]
            # check if the first parts of the address are both numeric
            if (addr_num1.isnumeric()) and (addr_num2.isnumeric()):
                if addr_num1 != addr_num2:
                    # if the nums dont match reduce their ratio 'score' by 12
                    f_tkn_ratio -= 12
                    f_ratio -= 12
                else:
                    f_tkn_ratio += 2
                    f_ratio += 2
            else:
                # then one address must contain a leading number and the other
                # doesnt, deduct 12
                f_tkn_ratio -= 12
                f_ratio -= 12

            # check the post codes match up
            # get the postcode from each address
            addr_1_pcode = re.findall(r'[A-Z]{1,2}[0-9R][0-9A-Z]? [0-9][A-Z]{2}', mod_addr_1)
            addr_2_pcode = re.findall(r'[A-Z]{1,2}[0-9R][0-9A-Z]? [0-9][A-Z]{2}', mod_addr_2)
            if addr_1_pcode != addr_2_pcode:
                # if the postcodes dont match reduce their ratio 'score' by 2
                f_tkn_ratio -= 2
                f_ratio -= 2
            else:
                f_tkn_ratio += 2
                f_ratio += 2
            pot_match = addr_match(addr_2, f_ratio, f_tkn_ratio, addr_1)
            # check if the address is a potential match
            # chck_ratio returns -1 if there's no match
            if chck_ratio(pot_match) != -1:
                # append the potential match to the matches list
                matches.append(pot_match)
            # print("DEBUG\n  ", mod_addr_1, "||", mod_addr_2, "\n\tratio: ",
            #     f_ratio, "tkn ratio: ", f_tkn_ratio, "tkn set ratio ",
            #     f_tkn_set_ratio, "partial ratio: ", f_partial_ratio)

        # choose the address from matches with the highest ratio
        # check there's at least one potential match to begin with
        if len(matches) > 0:
            match = matches[0]
            for addr in matches:
                if addr.f_tkn_ratio > match.f_tkn_ratio:
                    match = addr
            # now append the match and addr_1 to matched_addr list
            tier = chck_ratio(match)
            colour = set_console_col(tier)
            print(colour + "Found a match\n  " + addr_1 + " || " + match.addr + "\n\tratio: " + str(match.f_ratio), "tkn ratio: ", str(match.f_tkn_ratio), "tier:", str(tier) + Style.RESET_ALL)
            # matched_addr.append(match.addr)
            # matched_addr.append(addr_1)
            matched_addr.append(match)
    return matched_addr


# initialise multiprocessing lib ray
num_cpus = psutil.cpu_count(logical=True)
ray.init(num_cpus=num_cpus)

# parse the datasets
sa1_l = []
sa2_l = []
dataset = load_workbook('resources/Sample_Address_Data.xlsx')
worksheet = dataset.active
# iterate over Sample Address 1 addresses
for col_cells_sa1 in worksheet.iter_cols(min_col=1, max_col=1, min_row=2):
    for cell_sa1 in col_cells_sa1:
        sa1_l.append(cell_sa1.value)
# iterate over Sample Address 2 addresses
for col_cells_sa2 in worksheet.iter_cols(min_col=2, max_col=2, min_row=2):
    for cell_sa2 in col_cells_sa2:
        sa2_l.append(cell_sa2.value)

# store the lists in shared memory
ray.put(sa1_l)
ray.put(sa2_l)

# instantiate a matches_structure object
matches_data = match_structure()

# determine number of address before matching
num_addresses = len(sa1_l) + len(sa2_l)


# =========== PARALLEL SECTION ===========
# find matches in parallel using ray

# split sa1_l into (num_cpus) lists of approximate equaly length
sa1_l_split = split_list(sa1_l, num_cpus)

index = 0
results_id = []
for sa1_l_s in sa1_l_split:
    results_id.insert(index, find_matches_parallel.remote(sa1_l_s))
    index += 1

# =========== END PARALLEL ===============

# now that matches have been found from each process
# combine the results of each subprocess and populate the matches_data struct
# also populate matched_addr list for debug/performance metrics
results = ray.get(results_id)
print("--gathering results--")
matched_addr = []
for result in results:
    matched_addr.extend(result)
    # populate the matched_data structure
    for match in result:
        matches_data.insert_match(match)

# add the match to the matches dictionary
# matches_data.insert_match(match)

print("# Σ Addresses:", num_addresses)
print("matched", len(matched_addr)*2, "addresses")
print("# Σ unmatched addresses:", (num_addresses - len(matched_addr)*2))
# print a summary of matches in each tier
matches_data.print()


# print("\n\n\n\n")
# # print a list of unmatched addresses
# for addr in sa1_l:
#     if addr not in matched_addr:
#         print(addr)
